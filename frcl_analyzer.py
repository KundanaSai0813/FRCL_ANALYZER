import streamlit as st
import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from plotly.graph_objects import Figure, Scatter, Bar
import plotly.io as pio
from PIL import Image
import io
import os

st.set_page_config(page_title="FRCL Analyzer", layout="centered")

st.title("📊 FRCL Data Analyzer")
st.write("Upload your CSV file and get automated analysis with visual reports.")

uploaded_file = st.file_uploader("Upload CSV file", type=["csv"])

if uploaded_file is not None:
    try:
        data = pd.read_csv(uploaded_file)

        merged_data = data.merge(data, left_on='Barcode(AC)', right_on='Barcode(BC)', suffixes=('_AC', '_BC'))

        selected_data = merged_data[['Finishing Machine_AC', 'Production Date_AC', 'Barcode(AC)_AC', 'FRCL_AC', 'DeltaEZ1_BC']].rename(
            columns={
                'Finishing Machine_AC': 'Finishing Machine',
                'Production Date_AC': 'Production Date',
                'Barcode(AC)_AC': 'Barcode',
                'FRCL_AC': 'FRCL',
                'DeltaEZ1_BC': 'DeltaEZ1'
            }
        )

        selected_data['FRCL'] = pd.to_numeric(selected_data['FRCL'].str.replace(',', ''), errors='coerce')
        selected_data['DeltaEZ1'] = pd.to_numeric(selected_data['DeltaEZ1'], errors='coerce')

        filtered_data = selected_data[(selected_data['DeltaEZ1'] >= -3) & (selected_data['DeltaEZ1'] <= 3)]

        grouped = filtered_data.groupby('DeltaEZ1').agg({
            'Barcode': 'nunique',
            'FRCL': 'mean'
        }).reset_index().rename(columns={
            'Barcode': 'Number of Tires',
            'FRCL': 'Average FRCL'
        })

        X = sm.add_constant(grouped['Number of Tires'])
        y = grouped['Average FRCL']
        ols_model = sm.OLS(y, X).fit()
        grouped['OLS_Prediction'] = ols_model.predict(X)

        deltaez1_bins = np.arange(-3.0, 3.1, 0.5)
        bin_labels = [f'{deltaez1_bins[i]} to {deltaez1_bins[i + 1]}' for i in range(len(deltaez1_bins) - 1)]
        filtered_data['DeltaEZ1_Range'] = pd.cut(filtered_data['DeltaEZ1'], bins=deltaez1_bins, labels=bin_labels, right=False)
        range_count_df = filtered_data.groupby('DeltaEZ1_Range').agg({
            'Barcode': 'nunique'
        }).reset_index().rename(columns={'Barcode': 'Number of Tires'})

        # Plots
        fig1 = Figure()
        fig1.add_trace(Bar(x=grouped['DeltaEZ1'], y=grouped['Number of Tires'], name='Number of Tires', marker=dict(color='steelblue')))
        fig1.update_layout(title='Number of Tires by DeltaEZ1', xaxis_title='DeltaEZ1', yaxis_title='Number of Tires')

        fig2 = Figure()
        fig2.add_trace(Scatter(x=grouped['DeltaEZ1'], y=grouped['Average FRCL'], mode='markers',
                               name='Average FRCL', marker=dict(color='darkorange', size=10, line=dict(color='black', width=1))))
        fig2.add_trace(Scatter(x=grouped['DeltaEZ1'], y=grouped['OLS_Prediction'], mode='lines',
                               name='OLS Regression Line', line=dict(color='green', dash='dash')))
        fig2.update_layout(title='Average FRCL and OLS Regression by DeltaEZ1', xaxis_title='DeltaEZ1', yaxis_title='Average FRCL')

        fig3 = Figure()
        fig3.add_trace(Bar(x=range_count_df['DeltaEZ1_Range'], y=range_count_df['Number of Tires'], name='Number of Tires', marker=dict(color='green', opacity=0.7)))
        fig3.update_layout(title='Number of Tires in DeltaEZ1 Ranges (0.5 Interval)', xaxis_title='DeltaEZ1 Range', yaxis_title='Number of Tires')

        output_excel = "output_analysis.xlsx"

        # Save only the "Range Count" sheet
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            range_count_df.to_excel(writer, sheet_name='Range Count', index=False)

        # Load workbook and worksheet
        wb = load_workbook(output_excel)
        ws = wb['Range Count']

        # Apply formatting
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                if cell.row == 1:
                    cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

        # Function to insert Plotly plots into Excel
        def insert_plot(ws, fig, anchor_cell):
            img_bytes = pio.to_image(fig, format='png', width=800, height=400, engine='kaleido')
            img = Image.open(io.BytesIO(img_bytes))
            temp_img = "temp_plot.png"
            img.save(temp_img)
            xl_img = XLImage(temp_img)
            ws.add_image(xl_img, anchor_cell)
            os.remove(temp_img)

        # Insert plots into the Excel sheet
        insert_plot(ws, fig1, "D2")
        insert_plot(ws, fig2, "D22")
        insert_plot(ws, fig3, "D42")

        wb.save(output_excel)

        # Display on Streamlit
        st.success("✅ Analysis complete. See plots and download below.")

        st.plotly_chart(fig1, use_container_width=True)
        st.plotly_chart(fig2, use_container_width=True)
        st.plotly_chart(fig3, use_container_width=True)

        with open(output_excel, 'rb') as f:
            st.download_button("📥 Download Excel Report", f, file_name="FRCL_Analysis.xlsx")

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")
